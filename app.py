import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile

# Function to format the transcript with additional formatting logic
def format_transcript(input_file, output_path):
    # Read the uploaded CSV file
    data = pd.read_csv(input_file)
    
    # Prepare formatted conversation
    conversation = []
    for _, row in data.iterrows():
        start_time = row.get('start_time', 'Unknown Time')
        row_type = row.get('type', 'Unknown Type')
        if row_type not in ['debug', 'goto', 'knowledgeBase'] and pd.notna(row_type):
            if row_type == 'choice' and pd.notna(row.get('response')):
                buttons = row['response'].replace(',', ', ')
                conversation.append({
                    'START_TIME': start_time,
                    'TYPE': row_type,
                    'AGENT': f"BUTTONS DISPLAYED: {buttons}",
                    'USER': '',
                    'INTENT_MATCHED': ''
                })
            else:
                conversation.append({
                    'START_TIME': start_time,
                    'TYPE': row_type,
                    'AGENT': row.get('response', ''),
                    'USER': row.get('user_input', ''),
                    'INTENT_MATCHED': row.get('intent_matched', '')
                })

    # Create a DataFrame from the formatted data
    formatted_data = pd.DataFrame(conversation)

    # Use a temporary file for the Excel output
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
        # Save the DataFrame to Excel (explicitly specify the engine)
        formatted_data.to_excel(tmpfile.name, index=False, engine='openpyxl')
        tmpfile.close()

        # Re-open the file with openpyxl for formatting
        workbook = load_workbook(tmpfile.name)
        sheet = workbook.active
        pink_fill = PatternFill(start_color="FFD1DC", end_color="FFD1DC", fill_type="solid")
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Add formatting
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            agent_cell = row[2]
            user_cell = row[3]
            if agent_cell.value:
                agent_cell.fill = pink_fill
            if user_cell.value:
                user_cell.fill = grey_fill

        # Save the workbook again after applying formatting
        workbook.save(tmpfile.name)
        
        return tmpfile.name

# Streamlit App Interface
st.title("Voiceflow Transcript Formatter")

# File uploader
uploaded_file = st.file_uploader("Upload your transcript (CSV file)", type=["csv"])

# Input for the output file name
output_name = st.text_input("Enter a name for the output file (e.g., formatted_transcript.xlsx)")

# Button to process the transcript
if uploaded_file and output_name:
    if st.button("Process Transcript"):
        # Process the file and get the path to the temporary formatted file
        output_path = format_transcript(uploaded_file, f"/tmp/{output_name}")

        # Provide a download button for the formatted file
        with open(output_path, "rb") as f:
            st.download_button(
                label="Download Formatted Transcript",
                data=f,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
